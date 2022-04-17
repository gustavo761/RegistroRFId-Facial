from datetime import datetime
from logging import exception
from openpyxl import Workbook, load_workbook


meses = ['','ENERO','FEBRERO','MARZO','ABRIL','MAYO','JUNIO','JULIO','AGOSTO','SEPTIEMBRE','OCTUBRE','NOVIEMBRE','DICIEMBRE']


def reporteIndividual(datos):
    #poner dentro de un try
    now = datetime.now()
    pathPlantilla = 'plantillaReporte/plantillaIndividual.xlsx'
    pathGuardarReporte = f'plantillaReporte/{now.day}-{now.month}-{now.year} {now.hour}-{now.minute}-{now.second} Individual.xlsx'
    wb = load_workbook(pathPlantilla)
    ws = wb['Hoja1']
    ws.title = 'ReporteDeAsistencia'
    ws['C11'] = f"REPORTE DE ASISTENCIA DE {meses[int(datos[1])]} DE {int(datos[2])}"
    ws['B14'] = datos[5]
    ws['C14'] = datos[4]
    ws['E14'] = datos[0]
    try:
        ws['F14'] = datos[3]["numero"]
    except:
        ws['F14'] = "Número no registrado"
    fila = 17
    columna = 2
    for i in datos[6]:
        columna = 2
        for j in i:
            #sprint(i[j])
            ws.cell(row=fila,column=columna).value = i[j]
            columna += 1
        fila += 1
    
    wb.save(pathGuardarReporte)
    print("reporte individual")

def reporteGeneral(datos):
    now = datetime.now()
    pathPlantilla = 'plantillaReporte/plantillaGeneral.xlsx'
    pathGuardarReporte = f'plantillaReporte/{now.day}-{now.month}-{now.year} {now.hour}-{now.minute}-{now.second} General.xlsx'
    wb = load_workbook(pathPlantilla)
    ws = wb['Hoja1']
    ws['C11'] = f"REPORTE DE ASISTENCIA DE {meses[int(datos[0][1])]} DE {int(datos[0][2])}"
    ws.title = 'ReporteDeAsistencia'
    fila = 13
    columna = 2
    for usuarios in datos:   
        columna = 2 
        ws.cell(row=fila,column=columna).value = "TAG" 
        ws.cell(row=fila+1,column=columna).value = usuarios[5]
        ws.cell(row=fila,column=columna+1).value = "NOMBRES Y APELLIDOS" 
        ws.cell(row=fila+1,column=columna+1).value = usuarios[4]
        ws.cell(row=fila,column=columna+3).value = "CARNET" 
        ws.cell(row=fila+1,column=columna+3).value = usuarios[0]
        ws.cell(row=fila,column=columna+4).value = "CELULAR" 
        try:
            ws.cell(row=fila+1,column=columna+4).value = usuarios[3]["numero"]
        except:
            ws.cell(row=fila+1,column=columna+4).value = "Número no registrado"
        
        fila += 3
        ws.cell(row=fila,column=columna).value = "MODO REGISTRO" 
        ws.cell(row=fila,column=columna+1).value = "FECHA" 
        ws.cell(row=fila,column=columna+2).value = "HORA LLEGADA" 
        ws.cell(row=fila,column=columna+3).value = "HORA SALIDA" 
        ws.cell(row=fila,column=columna+4).value = "HORAS DE TRABAJO" 
        fila += 1

        for i in usuarios[6]:
            columna = 2
            for j in i:
                #print(i[j])
                ws.cell(row=fila,column=columna).value = i[j]
                columna += 1
            fila += 1
        fila += 5

    wb.save(pathGuardarReporte)
    print("Reporte general")

#reporteIndividual([7078362,11,2021])